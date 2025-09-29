#!/usr/bin/env python3
"""
Debug Excel column detection to see what's happening with column 16-36 detection
"""

import tempfile
import os
from openpyxl import load_workbook
import xlrd

def debug_excel_columns(file_path):
    """Debug column detection in an Excel file"""
    print(f"Debugging Excel file: {file_path}")
    
    # Try openpyxl first
    try:
        workbook = load_workbook(file_path, data_only=True)
        print(f"\n=== OPENPYXL (.xlsx) Analysis ===")
        
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            print(f"\nSheet: {sheet_name}")
            print(f"  max_row: {worksheet.max_row}")
            print(f"  max_column: {worksheet.max_column}")
            
            # Check actual data in first few rows to see column extent
            print(f"  Checking first 3 rows for actual data extent:")
            for row_num in range(1, min(4, worksheet.max_row + 1)):
                row_data = []
                actual_max_col = 0
                
                # Check way more columns than max_column suggests
                for col_num in range(1, 50):  # Check up to column 50
                    cell_value = worksheet.cell(row=row_num, column=col_num).value
                    if cell_value is not None and str(cell_value).strip():
                        actual_max_col = col_num
                        row_data.append(f"Col{col_num}:{str(cell_value)[:20]}")
                    elif col_num <= 40:  # Still show blanks up to col 40
                        row_data.append(f"Col{col_num}:BLANK")
                        
                print(f"    Row {row_num}: actual_max_col={actual_max_col}")
                # Show first 20 columns
                print(f"    First 20 cols: {' | '.join(row_data[:20])}")
                if len(row_data) > 20:
                    print(f"    Cols 21-40: {' | '.join(row_data[20:40])}")
                    
    except Exception as e:
        print(f"openpyxl failed: {e}")
    
    # Try xlrd for .xls
    try:
        workbook = xlrd.open_workbook(file_path)
        print(f"\n=== XLRD (.xls) Analysis ===")
        
        for sheet_index in range(workbook.nsheets):
            sheet = workbook.sheet_by_index(sheet_index)
            print(f"\nSheet: {sheet.name}")
            print(f"  nrows: {sheet.nrows}")
            print(f"  ncols: {sheet.ncols}")
            
            # Check actual data extent
            print(f"  Checking first 3 rows for actual data extent:")
            for row_index in range(min(3, sheet.nrows)):
                row_data = []
                actual_max_col = 0
                
                # Check more columns than ncols suggests
                for col_index in range(min(50, sheet.ncols + 20)):
                    try:
                        cell_value = sheet.cell_value(row_index, col_index)
                        if cell_value is not None and str(cell_value).strip():
                            actual_max_col = col_index
                            row_data.append(f"Col{col_index}:{str(cell_value)[:20]}")
                        elif col_index < 40:
                            row_data.append(f"Col{col_index}:BLANK")
                    except:
                        break
                        
                print(f"    Row {row_index}: actual_max_col={actual_max_col}")
                print(f"    First 20 cols: {' | '.join(row_data[:20])}")
                if len(row_data) > 20:
                    print(f"    Cols 21-40: {' | '.join(row_data[20:40])}")
                    
    except Exception as e:
        print(f"xlrd failed: {e}")

if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1:
        debug_excel_columns(sys.argv[1])
    else:
        print("Usage: python debug_excel_columns.py <excel_file_path>")