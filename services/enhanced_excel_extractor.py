#!/usr/bin/env python3
"""
Enhanced Excel Extractor with Data Cleaning and Normalization
Preprocesses Excel files to create clean, consistent format for extraction functions
"""

import sys
import json
import base64
import io
import tempfile
import os
import re
from typing import List, Dict, Any, Optional

try:
    import pandas as pd
    from openpyxl import load_workbook
    import xlrd
except ImportError as e:
    print(f"Error: Missing required library: {e}", file=sys.stderr)
    sys.exit(1)

class ExcelPreprocessor:
    """Enhanced Excel preprocessor that cleans and normalizes data"""
    
    def __init__(self):
        self.debug = False  # Disable debug output for production use
        
    def clean_cell_value(self, value: Any) -> str:
        """Clean and normalize a single cell value"""
        if value is None or pd.isna(value):
            return ""
            
        # Convert to string
        text = str(value).strip()
        
        # Remove common Excel formatting artifacts
        text = text.replace('\x00', '')  # Remove null characters
        text = re.sub(r'\s+', ' ', text)  # Normalize whitespace
        text = text.replace('\r\n', ' ').replace('\r', ' ').replace('\n', ' ')  # Handle line breaks
        
        # Clean up common Excel number formatting
        if text.endswith('.0') and text[:-2].isdigit():
            text = text[:-2]  # Remove .0 from whole numbers
            
        return text.strip()
    
    def detect_header_row(self, rows: List[List[str]]) -> int:
        """Detect which row contains the actual headers"""
        if not rows:
            return 0
            
        # Look for the row with the most non-empty, text-heavy cells
        best_row_idx = 0
        best_score = 0
        
        for i, row in enumerate(rows[:10]):  # Check first 10 rows
            if not row:
                continue
                
            score = 0
            non_empty_count = 0
            
            for cell in row:
                if cell and cell.strip():
                    non_empty_count += 1
                    # Prefer text over numbers for headers
                    if len(cell) > 5 and not cell.isdigit():
                        score += 2
                    elif len(cell) > 1:
                        score += 1
                        
            # Normalize score by row length
            if len(row) > 0:
                score = (score * non_empty_count) / len(row)
                
            if score > best_score and non_empty_count > 3:  # Minimum 3 headers
                best_score = score
                best_row_idx = i
                
        return best_row_idx
    
    def merge_multiline_headers(self, rows: List[List[str]], header_row_idx: int) -> List[str]:
        """Merge multi-line headers into single coherent headers"""
        if header_row_idx >= len(rows):
            return []
            
        base_headers = rows[header_row_idx]
        merged_headers = base_headers.copy()
        
        # Look for continuation lines after the main header
        continuation_rows = []
        for i in range(header_row_idx + 1, min(header_row_idx + 5, len(rows))):
            if i < len(rows):
                next_row = rows[i]
                
                # Check if this looks like a continuation (fewer columns, no numeric data)
                if (len(next_row) <= len(base_headers) * 0.7 and  # Fewer columns
                    any(len(cell) > 10 for cell in next_row if cell) and  # Has substantial text
                    not any(cell.replace('.', '').replace('-', '').isdigit() 
                           for cell in next_row[:5] if cell)):  # No leading numbers
                    
                    continuation_rows.append(next_row)
                else:
                    break  # Stop at first data-like row
        
        if self.debug:
            print(f"DEBUG - Found {len(continuation_rows)} continuation rows for headers", file=sys.stderr)
        
        # Merge continuation text into headers
        for cont_row in continuation_rows:
            for i, cont_text in enumerate(cont_row):
                if cont_text and cont_text.strip():
                    if i < len(merged_headers):
                        # Check if the base header looks incomplete
                        base_header = merged_headers[i]
                        if (base_header.endswith('(') or base_header.endswith('And') or 
                            base_header.endswith('Subject To') or base_header.endswith('pa') or
                            base_header.endswith('for') or len(base_header) > 50):
                            
                            # Merge the continuation
                            merged_headers[i] = (base_header + " " + cont_text).strip()
                            
                            if self.debug:
                                print(f"DEBUG - Merged header {i}: '{base_header}' + '{cont_text}'", file=sys.stderr)
        
        return merged_headers
    
    def normalize_column_count(self, headers: List[str], target_count: Optional[int] = None) -> List[str]:
        """Normalize column count and fill missing headers"""
        if not target_count:
            target_count = len(headers)
            
        normalized = headers.copy()
        
        # Fill missing headers with positional names
        while len(normalized) < target_count:
            col_num = len(normalized) + 1
            normalized.append(f"Column_{col_num}")
            
        # Truncate if too many
        if len(normalized) > target_count:
            normalized = normalized[:target_count]
            
        return normalized
    
    def extract_excel_text_enhanced(self, file_content: bytes, file_name: str) -> str:
        """Enhanced Excel extraction with preprocessing and cleaning"""
        try:
            text_parts = []
            
            if self.debug:
                print(f"DEBUG - Processing Excel file: {file_name}", file=sys.stderr)
            
            # Try modern Excel format first (.xlsx)
            workbook_data = None
            try:
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                    tmp_file.write(file_content)
                    tmp_file.flush()
                    
                    workbook = load_workbook(tmp_file.name, data_only=True)
                    workbook_data = {}
                    
                    for sheet_name in workbook.sheetnames:
                        worksheet = workbook[sheet_name]
                        sheet_data = []
                        
                        for row in worksheet.iter_rows(values_only=True, max_row=100):  # Limit to first 100 rows for headers
                            row_data = [self.clean_cell_value(cell) for cell in row]
                            if any(cell for cell in row_data):  # Only keep non-empty rows
                                sheet_data.append(row_data)
                        
                        workbook_data[sheet_name] = sheet_data
                    
                    os.unlink(tmp_file.name)
                    
            except Exception as xlsx_error:
                if self.debug:
                    print(f"DEBUG - XLSX failed, trying XLS: {xlsx_error}", file=sys.stderr)
                
                # Fall back to older Excel format (.xls)
                try:
                    with tempfile.NamedTemporaryFile(delete=False, suffix='.xls') as tmp_file:
                        tmp_file.write(file_content)
                        tmp_file.flush()
                        
                        workbook = xlrd.open_workbook(tmp_file.name)
                        workbook_data = {}
                        
                        for sheet_index in range(workbook.nsheets):
                            sheet = workbook.sheet_by_index(sheet_index)
                            sheet_name = sheet.name
                            sheet_data = []
                            
                            for row_index in range(min(sheet.nrows, 100)):  # Limit for headers
                                row_values = sheet.row_values(row_index)
                                row_data = [self.clean_cell_value(val) for val in row_values]
                                if any(cell for cell in row_data):
                                    sheet_data.append(row_data)
                            
                            workbook_data[sheet_name] = sheet_data
                        
                        os.unlink(tmp_file.name)
                        
                except Exception as xls_error:
                    if self.debug:
                        print(f"DEBUG - XLS failed, trying pandas: {xls_error}", file=sys.stderr)
                    
                    # Final fallback using pandas
                    excel_data = pd.read_excel(io.BytesIO(file_content), sheet_name=None, nrows=100)
                    workbook_data = {}
                    
                    for sheet_name, df in excel_data.items():
                        sheet_data = []
                        
                        # Add headers
                        headers = [self.clean_cell_value(h) for h in df.columns.tolist()]
                        sheet_data.append(headers)
                        
                        # Add data rows
                        for _, row in df.iterrows():
                            row_data = [self.clean_cell_value(val) for val in row.values]
                            sheet_data.append(row_data)
                        
                        workbook_data[sheet_name] = sheet_data
            
            if not workbook_data:
                raise Exception("Could not read Excel file with any method")
            
            # Process each sheet with enhanced logic
            for sheet_name, raw_rows in workbook_data.items():
                if not raw_rows:
                    continue
                    
                if self.debug:
                    print(f"DEBUG - Processing sheet '{sheet_name}' with {len(raw_rows)} rows", file=sys.stderr)
                
                text_parts.append(f"=== Sheet: {sheet_name} ===")
                
                # Detect the real header row
                header_row_idx = self.detect_header_row(raw_rows)
                
                if self.debug:
                    print(f"DEBUG - Detected header at row {header_row_idx}", file=sys.stderr)
                
                # Extract and merge multi-line headers
                headers = self.merge_multiline_headers(raw_rows, header_row_idx)
                
                if self.debug:
                    print(f"DEBUG - Final headers: {len(headers)} columns", file=sys.stderr)
                    print(f"DEBUG - Sample headers: {headers[:3]} ... {headers[-3:] if len(headers) > 6 else headers}", file=sys.stderr)
                
                # Add the cleaned header row
                text_parts.append("\t".join(headers))
                
                # Add a few sample data rows for context (but functions will focus on headers)
                data_rows_added = 0
                for i in range(header_row_idx + 1, len(raw_rows)):
                    if data_rows_added >= 3:  # Limit sample data
                        break
                        
                    row = raw_rows[i]
                    if any(cell for cell in row):  # Non-empty row
                        # Normalize row length to match headers
                        normalized_row = row[:len(headers)] + [""] * max(0, len(headers) - len(row))
                        text_parts.append("\t".join(normalized_row))
                        data_rows_added += 1
            
            result = "\n".join(text_parts)
            
            if self.debug:
                print(f"DEBUG - Final extracted content: {len(result)} characters", file=sys.stderr)
                print(f"DEBUG - Tab count: {result.count(chr(9))} tabs", file=sys.stderr)
            
            return result
            
        except Exception as e:
            if self.debug:
                print(f"DEBUG - Excel extraction failed: {str(e)}", file=sys.stderr)
            raise Exception(f"Enhanced Excel extraction failed: {str(e)}")

def main():
    """Main function for standalone usage"""
    if len(sys.argv) < 2:
        print("Usage: python enhanced_excel_extractor.py <input_json>")
        sys.exit(1)
    
    # Read input JSON
    input_data = json.loads(sys.argv[1])
    
    documents = input_data.get('documents', [])
    preprocessor = ExcelPreprocessor()
    
    results = []
    for doc in documents:
        try:
            # Decode base64 content
            file_content = base64.b64decode(doc['content'])
            file_name = doc.get('name', 'unknown.xlsx')
            
            # Extract text using enhanced method
            extracted_text = preprocessor.extract_excel_text_enhanced(file_content, file_name)
            
            results.append({
                'document_id': doc.get('id'),
                'extracted_content': extracted_text,
                'status': 'success'
            })
            
        except Exception as e:
            results.append({
                'document_id': doc.get('id'),
                'error': str(e),
                'status': 'error'
            })
    
    print(json.dumps(results, indent=2))

if __name__ == "__main__":
    main()